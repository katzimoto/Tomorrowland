"""XLSX text extractor."""

from __future__ import annotations

import contextlib
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from services.extraction.base import ExtractionResult


class XlsxExtractor:
    """Extract text from Excel .xlsx files using openpyxl."""

    def extract(self, path: Path) -> ExtractionResult:
        """Return concatenated text from all cells.

        Uses ``data_only=True`` so formula cells yield their last computed
        value rather than the formula string, giving translatable content.
        Uses ``read_only=True`` for memory efficiency on large sheets.
        ``wb.close()`` is always called via ``finally`` so file handles are
        released even when an exception occurs mid-iteration.
        """
        wb = None
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            texts: list[str] = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            texts.append(str(cell.value))
            return ExtractionResult(text="\n".join(texts))
        except (OSError, KeyError, ValueError, InvalidFileException, zipfile.BadZipFile):
            return ExtractionResult(text="")
        except Exception:  # noqa: BLE001
            return ExtractionResult(text="")
        finally:
            if wb is not None:
                with contextlib.suppress(Exception):
                    wb.close()
