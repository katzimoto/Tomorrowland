"""XLSX → per-sheet grid preview renderer.

Spreadsheets paginate poorly through PDF (wide sheets shred into fragments),
so instead of converting to PDF we emit one JSON grid artifact per sheet and
render it with a dedicated SheetViewer that has real sheet tabs. A preview
shows the first ``max_rows`` × ``max_cols`` cells; the full sheet stays
available via download. Uses ``data_only=True`` so formula cells yield their
last computed value.
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class RenderedSheets:
    """Sheet-grid renderer output consumed by the render orchestrator."""

    # artifact_id -> (relative filename, content type, bytes)
    artifacts: dict[str, tuple[str, str, bytes]]
    # navigation items: {index, label, artifact_id}
    sheets: list[dict[str, Any]]
    truncated: bool


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def render_sheets(
    source_path: Path,
    *,
    max_rows: int,
    max_cols: int,
) -> RenderedSheets:
    """Render each worksheet into a capped JSON grid artifact."""
    wb = None
    try:
        wb = load_workbook(str(source_path), read_only=True, data_only=True)
        artifacts: dict[str, tuple[str, str, bytes]] = {}
        sheets: list[dict[str, Any]] = []
        any_truncated = False

        for index, ws in enumerate(wb.worksheets):
            rows: list[list[str]] = []
            row_truncated = False
            col_truncated = False
            for r, row in enumerate(ws.iter_rows(values_only=True)):
                if r >= max_rows:
                    row_truncated = True
                    break
                cells = [_cell_to_str(v) for v in row[:max_cols]]
                if len(row) > max_cols:
                    col_truncated = True
                rows.append(cells)

            truncated = row_truncated or col_truncated
            any_truncated = any_truncated or truncated
            artifact_id = f"sheet-{index}"
            grid = {
                "name": ws.title,
                "rows": rows,
                "truncated": {"rows": row_truncated, "cols": col_truncated},
            }
            artifacts[artifact_id] = (
                f"{artifact_id}.json",
                "application/json",
                json.dumps(grid).encode("utf-8"),
            )
            sheets.append({"index": index, "label": ws.title, "artifact_id": artifact_id})

        return RenderedSheets(artifacts=artifacts, sheets=sheets, truncated=any_truncated)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                logger.debug("failed to close workbook", exc_info=True)
